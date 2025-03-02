import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from services import datamanager
import re

output_file = 'DQMatrixOut.xlsx'

def remove_digits(text):
    return re.sub(r"\d", "", text)

def check_for_merge(worksheet, match_start, length):
    if length > 0:
        #todo initiate merge
        print(worksheet.cell(row=1, column=match_start).value)
        print(worksheet.cell(row=1, column=match_start + length).value)

def generate_dynamic_attribute_rows(data, name_cols, phone_cols, num_cols, address_cols, city_cols, state_cols, postal_cols, email_cols):
    # generate name columns
    for i in range(0, len(name_cols)):
        data['name cols' + str(i)] = [name_cols[i], 'TRUE', '', '', 'TRUE', '', 'TRUE', '', '', '', '', '']

    # generate number columns
    for i in range(0, len(num_cols) + len(phone_cols)):
        if i < len(num_cols):
            data['number cols' + str(i)] = [num_cols[i], 'TRUE', 'TRUE', '', '', '', 'TRUE', '', '', '', 'TRUE', '']
        else:
            # generate phone number columns
            for j in range(0, len(phone_cols)):
                data['number cols' + str(i)] = [phone_cols[j], 'TRUE', 'TRUE', '', '', '', 'TRUE', '', '', '', 'TRUE',
                                                '']
            break
    # generate address columns
    for i in range(0, len(address_cols) + len(city_cols) + len(state_cols) + len(postal_cols)):
        print()
        if i < len(address_cols):
            data['addr_other_cols' + str(i)] = [address_cols[i], 'TRUE', '', '', '', '', 'TRUE', '', '', '', '', 'TRUE']
        elif i < len(address_cols) + len(city_cols):
            # generate city columns
            for j in range(0, len(city_cols)):
                data['addr_other_cols' + str(i)] = [city_cols[j], 'TRUE', '', '', '', '', 'TRUE', '', '', '', '','TRUE']
        elif i < len(address_cols) + len(city_cols) + len(state_cols):
            # generate state columns
            for j in range(0, len(state_cols)):
                data['addr_other_cols' + str(i)] = [state_cols[j], '', '', '', '', '', 'TRUE', '', 'TRUE', '', '','']
        elif i < len(address_cols) + len(city_cols) + len(state_cols) + len(postal_cols):
            # generate postal columns
            for j in range(0, len(postal_cols)):
                data['addr_other_cols' + str(i)] = [postal_cols[j], 'TRUE', '', '', 'TRUE', '', 'TRUE', '', '', '', '','']
            break
    # generate email columns
    for i in range(0, len(email_cols)):
        data['email_cols' + str(i)] = [email_cols[j], 'TRUE', '', 'TRUE', '', '', 'TRUE', '', '', '', 'TRUE', '']
    return data

def merge_engine(worksheet):
    i = 2
    j = 1
    length = 0
    match_start = 0
    while worksheet.cell(row=1, column=i).value is not None:
        # use 2 pointer notation to iterate over column headers
        prev_value = remove_digits(worksheet.cell(row=1, column=j).value)
        curr_value = remove_digits(worksheet.cell(row=1, column=i).value)

        if prev_value == curr_value:  # if a match is found
            i = i + 1  # move runner + 1 to check next value
            match_start = j  # update match start
            length = length + 1  # increment match length
        else:  # if a match is not found or had ended
            check_for_merge(worksheet, match_start, length)  # check to see if the length is > 1. If so initiate a merge
            j = i
            i = i + 1
            length = 0

    # check for merges after iteration
    check_for_merge(worksheet, match_start, length)


def build_excel_sheet(sheet_name, name_cols, num_cols, phone_cols, address_cols, city_cols, state_cols, postal_cols, email_cols):
    # build data dic for static columns
    data = {
        'Workflow': ['Rule Num', 'DQ-1', 'DQ-2', 'DQ-3', 'DQ-4', 'DQ-5', 'DQ-6', 'DQ-7', 'DQ-8', 'DQ-9', 'DQ-10',
                     'DQ-11'],
        'transformation type': ['DQ rule name', 'Trim', 'Remove Non-numeric characters', 'Lower case', 'Upper case',
                                'Remove numeric characters', 'Standardize blank values to Null',
                                'Remove Special Characters', 'Standardize State codes', 'Standardize Country Codes',
                                'Remove White Space', 'Title case'],
    }
    generate_dynamic_attribute_rows(data, name_cols, phone_cols, num_cols, address_cols, city_cols, state_cols, postal_cols, email_cols)

    # Create DataFrame
    df = pd.DataFrame(data)

    # Save to Excel file
    datamanager.write_to_excel(df, output_file, sheet_name)

    workbook = datamanager.load_excel(output_file)
    worksheet = workbook[sheet_name]

    merge_engine(worksheet) # check to see if header columns need merged



def build_definition_sheet():
    # Data for the workflow table
    data = {
        'Build Status': ['TRUE', 'TRUE', 'TRUE', 'TRUE', 'TRUE', 'TRUE', 'TRUE', 'TRUE',
                         'TRUE', 'TRUE', 'TRUE'],
        'Rule Number': ['DQ-1', 'DQ-2', 'DQ-3', 'DQ-4', 'DQ-5', 'DQ-6', 'DQ-7', 'DQ-8', 'DQ-9', 'DQ-10', 'DQ-11'],
        'DQ rule name': ['Trim', 'Remove Non-numeric characters', 'Lower case', 'Upper case',
                         'Remove numeric characters', 'Standardize blank values to Null', 'Remove Special Characters',
                         'Standardize State codes', 'Standardize Country Codes', 'Remove White Space', 'Title case'],
        'Description': [
            'left and right trim to remove leading and trailing spaces',
            'Removes any character from a string that is not 0-9',
            'Converts all characters to lower case characters',
            'Converts all characters to upper case characters',
            'Removes characters 0-9',
            'Null <blank> or padded space values',
            'Removes any non-standard characters from string',
            'Convert any US state values to Alpha-2 state codes',
            'Converts any country values to Alpha-3 country codes',
            'Removes any white space characters from a String',
            'Converts words in string to have a capital at the beginning'
        ],
        'Before Standardization': ['"John" ', '(987) 654-3210', '"John Doe"', '"John Doe"', '"John Doe 3"', '-',
                                   '"John @#$% Doe"', '"Ohio"', '"Canada"', 'kyle sinko@infoverity.com', '"billy bob"'],
        'After Standardization': ['"JOHN"', '9876543210', 'john doe', 'JOHN DOE', 'john doe', 'NULL', 'john doe', 'OH',
                                  'CAN', 'kylesinko@infoverity.com', 'Billy Bob'],

        'Query Name': ['dq_trim', 'dq_rm_non_numbers', 'dq_to_lower', 'dq_to_upper', 'dq_rm_numbers',
                       'dq_blank_to_null', 'dq_rm_special_char', 'dq_stdz_state', 'dq_stdz_country',
                       'dq_rm_white_space', 'dq_to_title_case'],
        'Notes': ['', '', '', '', '', '', '', '', '', '', '']
    }

    # Create DataFrame
    df = pd.DataFrame(data)
    datamanager.write_to_excel(df, output_file, 'DQ Rule List')

    workbook = datamanager.load_excel(output_file)
    worksheet = workbook['DQ Rule List']  # Use the sheet name from write_to_excel

    # Define colors and styles
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for header row
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')  # Center alignment for header row
    left_alignment = Alignment(horizontal='left')  # Left alignment for data cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # Thin border for all sides

    # Apply orange background, bold font, and center alignment to the header row (row 1)
    for cell in worksheet[1]:
        cell.fill = orange_fill
        cell.font = bold_font
        cell.alignment = center_alignment

    # Apply left alignment and borders to all data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.alignment = left_alignment
            cell.border = thin_border

    # Adjust column widths for readability
    column_widths = []
    for row in worksheet.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths.append(len(str(cell.value)))
    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[chr(64 + i)].width = column_width + 2  # Add padding

    # Save the workbook with formatting
    datamanager.save_excel_formating(workbook, output_file)
